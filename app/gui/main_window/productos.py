import logging
from PyQt5.QtCore import Qt, QRect, QTimer
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QFormLayout,
    QHBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox,
    QTableWidget, QTableWidgetItem, QMessageBox, QTabWidget,
    QRadioButton, QButtonGroup, QSpinBox, QInputDialog, QMenu, QFileDialog,
    QCheckBox, QStyle, QHeaderView, QDialog, QDoubleSpinBox,QCompleter
)
from app.gui.dialogs import _draw_barcode_label
from barcode.writer import ImageWriter
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from app.utils_timing import measure
from PyQt5.QtGui import QPainter, QPixmap, QIcon, QMouseEvent, QFont, QFontMetrics
from app.gui.common import BASE_ICONS_PATH, MIN_BTN_HEIGHT, ICON_SIZE, icon, _safe_viewport, _mouse_release_event_type, _checked_states

import pandas as pd
import barcode
import os
import re
from io import BytesIO

logger = logging.getLogger(__name__)


# Dependencias del dominio / helpers que usan tus métodos de productos:
from app.models import Producto
from app.repository import prod_repo
from app.gui.qt_helpers import freeze_table
# Si al ejecutar obtienes NameError: Icon -> descomenta la siguiente línea:
# from app.gui.qt_helpers import Icon

class ProductosMixin:
  
    #Métodos de la pestaña Productos extraídos de MainWindow.
    def tab_productos(self):
        w = QWidget()
        layout = QVBoxLayout()

        # Set para mantener IDs de productos seleccionados entre búsquedas
        self._selected_product_ids = set()
        # Log de cambios masivos (se acumula durante la sesión)
        self._product_change_log = []

        # Live search
        self.input_buscar = QLineEdit()
        self.input_buscar.setPlaceholderText('Buscar: código, nombre, categoría o precio (separar con coma/espacio)')
        # Debounce: espera 250ms tras dejar de escribir antes de buscar
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(250)
        self._search_timer.timeout.connect(self._do_buscar_productos)
        self.input_buscar.textChanged.connect(self._on_buscar_text_changed)
        layout.addWidget(self.input_buscar)

        # —— Estilos para el Live Search —— 
        font = self.input_buscar.font()
        font.setPointSize(12)
        font.setBold(True)
        self.input_buscar.setFont(font)
        self.input_buscar.setMinimumHeight(30)
        self.input_buscar.setStyleSheet("padding: 4px;")

        # Completer: autocompletado con sugerencias "CÓDIGO - NOMBRE"
        try:
            from app.gui.ventas_helpers import build_product_completer
            comp_prod, model_prod = build_product_completer(self.session, self)
            self._completer_productos = comp_prod
            self._completer_productos_model = model_prod
            self.input_buscar.setCompleter(comp_prod)

            # Al seleccionar del completer, quedarse solo con el código
            comp_prod.activated.connect(
                lambda s: self.input_buscar.setText(str(s).split(" - ")[0].strip())
            )

            # Debounce para actualizar filtro del completer
            from PyQt5.QtCore import QTimer as _QTimer
            self._completer_prod_debounce = _QTimer(self)
            self._completer_prod_debounce.setSingleShot(True)
            self._completer_prod_debounce.setInterval(150)
            self._completer_prod_debounce.timeout.connect(self._update_completer_productos)
            self.input_buscar.textChanged.connect(lambda t: self._completer_prod_debounce.start())
        except Exception:
            pass

        # Formulario
        form = QFormLayout()
        self.input_codigo    = QLineEdit()
        self.input_nombre    = QLineEdit()
        self.input_precio    = QLineEdit()
        self.input_categoria = QLineEdit()
        form.addRow('Código de Barra:',     self.input_codigo)
        form.addRow('Nombre:',              self.input_nombre)
        form.addRow('Precio:',              self.input_precio)
        form.addRow('Categoría (opcional):',self.input_categoria)

        # Encadenar foco: Código→Nombre→Precio→Categoría→Agregar
        self.input_codigo.returnPressed.connect(lambda: self.input_nombre.setFocus())
        self.input_nombre.returnPressed.connect(lambda: self.input_precio.setFocus())
        self.input_precio.returnPressed.connect(lambda: self.input_categoria.setFocus())
        self.input_categoria.returnPressed.connect(self.agregar_producto)

        # Forzar MAYÚSCULAS en nombre y categoría
        self.input_nombre.textChanged.connect(
            lambda t: self.input_nombre.setText(t.upper()) if t != t.upper() else None)
        self.input_categoria.textChanged.connect(
            lambda t: self.input_categoria.setText(t.upper()) if t != t.upper() else None)

        # —— Botones —— 
        btn_a   = QPushButton('Agregar/Actualizar'); btn_a.setIcon(icon('add.svg'));    btn_a.setToolTip('Agregar/Actualizar');    btn_a.clicked.connect(self.agregar_producto)
        btn_d   = QPushButton('Eliminar seleccionados'); btn_d.setIcon(icon('delete.svg')); btn_d.setToolTip('Eliminar seleccionados'); btn_d.clicked.connect(self.eliminar_productos)
        btn_u   = QPushButton('Deshacer'); btn_u.setIcon(icon('undo.svg'));   btn_u.setToolTip('Deshacer');               btn_u.clicked.connect(self.deshacer)
        btn_p   = QPushButton('Imprimir códigos'); btn_p.setIcon(icon('print.svg'));  btn_p.setToolTip('Imprimir códigos');       btn_p.clicked.connect(self.imprimir_codigos)
        btn_m   = QPushButton('Edición masiva'); btn_m.setIcon(icon('edit.svg'));   btn_m.setToolTip('Editar precio, nombre o categoría de los seleccionados'); btn_m.clicked.connect(self.editar_precios_masivos)
        btn_imp = QPushButton('Importar desde Excel'); btn_imp.setIcon(icon('import.svg'));btn_imp.setToolTip('Importar desde Excel'); btn_imp.clicked.connect(self.importar_productos)
        btn_exp = QPushButton('Exportar a Excel'); btn_exp.setIcon(icon('export.svg'));btn_exp.setToolTip('Exportar a Excel');      btn_exp.clicked.connect(self.exportar_productos)
        btn_cambios = QPushButton('Últimos cambios'); btn_cambios.setIcon(icon('history.svg'));btn_cambios.setToolTip('Ver historial de ediciones masivas');btn_cambios.setMinimumHeight(MIN_BTN_HEIGHT);btn_cambios.setIconSize(ICON_SIZE);btn_cambios.clicked.connect(self._abrir_ultimos_cambios)
        btn_sin_ventas = QPushButton('Sin movimiento'); btn_sin_ventas.setIcon(icon('list.svg'));btn_sin_ventas.setToolTip('Productos sin ventas en 90 días');btn_sin_ventas.setMinimumHeight(MIN_BTN_HEIGHT);btn_sin_ventas.setIconSize(ICON_SIZE);btn_sin_ventas.clicked.connect(self._ver_productos_sin_ventas)
        self.lbl_paginacion = QLabel("")

        hb = QHBoxLayout()
        for btn in (btn_a, btn_d, btn_u, btn_m, btn_imp, btn_exp, btn_p, btn_cambios, btn_sin_ventas):
            btn.setMinimumHeight(MIN_BTN_HEIGHT)
            btn.setIconSize(ICON_SIZE)
            hb.addWidget(btn)

        form.addRow(hb)
        layout.addLayout(form)

        # Tabla de productos (sin ID visible)
        self.table_productos = QTableWidget()
        self.table_productos.setColumnCount(6)
        self.table_productos.setHorizontalHeaderLabels(
            ['Sel','ID','Código', 'Nombre', 'Precio', 'Categoría']
        )
        
        self.table_productos.setFont(QFont("Arial", 11))
        self.table_productos.setSortingEnabled(True)
        self.table_productos.setSelectionMode(QTableWidget.MultiSelection)
        self.table_productos.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_productos.customContextMenuRequested.connect(self.menu_contexto_productos)
        self.table_productos.verticalHeader().setVisible(False)

        # Ajustar ancho columnas: Interactive para que el usuario pueda redimensionar
        hdr = self.table_productos.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        hdr.setStretchLastSection(True)
        self.table_productos.setColumnWidth(0, 28)   # Sel
        self.table_productos.setColumnWidth(1, 48)   # ID
        self.table_productos.setColumnWidth(2, 200)  # Código
        self.table_productos.setColumnWidth(3, 300)  # Nombre
        self.table_productos.setColumnWidth(4, 100)  # Precio

        layout.addWidget(self.table_productos)
        layout.addWidget(self.lbl_paginacion)

        self.table_productos.itemChanged.connect(self._on_producto_item_changed)
        self.table_productos.cellClicked.connect(self.cargar_producto)
        self.table_productos.cellDoubleClicked.connect(self.cargar_producto)

        w.setLayout(layout)

        # ESC para limpiar todos los campos del formulario y búsqueda
        from PyQt5.QtWidgets import QShortcut
        from PyQt5.QtGui import QKeySequence
        esc_shortcut = QShortcut(QKeySequence(Qt.Key_Escape), w)
        esc_shortcut.activated.connect(self._limpiar_campos_productos)

        # ---- Instalar filtro seguro ----
        try:
            vp = self.table_productos.viewport()
            if vp is not None:
                vp.installEventFilter(self)
        except Exception as e:
            logger.error("Error instalando filtro en Productos: %s", e)

        def _safe_remove_filter():
            try:
                tbl = getattr(self, "table_productos", None)
                if tbl is not None:
                    vp = getattr(tbl, 'viewport', lambda: None)()
                    if vp is not None:
                        vp.removeEventFilter(self)
            except Exception:
                pass
        self.table_productos.destroyed.connect(_safe_remove_filter)

        self.refrescar_productos()
        return w

        
    def _on_productos_cell_clicked(self, row, col):
        if col != 0:
            return
        it = self.table_productos.item(row, 0)
        if it is None:
            return
        it.setCheckState(Qt.Unchecked if it.checkState() == Qt.Checked else Qt.Checked)
    def menu_contexto_productos(self, pos):
        menu   = QMenu()
        act_del= menu.addAction('Eliminar seleccionado')
        action = menu.exec_(self.table_productos.viewport().mapToGlobal(pos))
        if action == act_del:
            idx = self.table_productos.indexAt(pos)
            if idx.isValid():
                pid = int(self.table_productos.item(idx.row(),1).text())
                p   = self.session.query(Producto).get(pid)
                if p:
                    datos = {'codigo_barra':p.codigo_barra,'nombre':p.nombre,
                            'precio':p.precio,'categoria':p.categoria}
                    self.history.append(('del',[datos]))
                    self.session.delete(p); self.session.commit()
                    self.statusBar().showMessage('Producto eliminado',3000)
                    self.refrescar_productos()
    def agregar_producto(self):
        with measure("actualizar_total"):
            c  = self.input_codigo.text().strip()
            n  = self.input_nombre.text().strip().upper()
            try: pr = float(self.input_precio.text())
            except: return QMessageBox.warning(self,'Error','Precio inválido')
            ca = self.input_categoria.text().strip().upper() or None

            editing_id = getattr(self, '_editing_product_id', None)
            existe = self.session.query(Producto).filter_by(codigo_barra=c).first()

            if existe and editing_id == existe.id:
                # Modo edición: actualizar producto existente
                datos_viejos = {'nombre': existe.nombre, 'precio': existe.precio, 'categoria': existe.categoria}
                existe.nombre = n
                existe.precio = pr
                existe.categoria = ca
                self.session.commit()
                self._sync_push("producto", existe)
                # Log de cambios por campo
                datos_nuevos = {'nombre': n, 'precio': pr, 'categoria': ca}
                for campo_k in ('nombre', 'precio', 'categoria'):
                    old_v = datos_viejos[campo_k]
                    new_v = datos_nuevos[campo_k]
                    if str(old_v or '') != str(new_v or ''):
                        self._log_product_change(existe, campo_k, old_v, new_v, 'Formulario - Actualizar')
                self.history.append(('edit', datos_viejos, datos_nuevos))
                self.statusBar().showMessage('Producto actualizado', 3000)
                self._beep_ok()
                self._editing_product_id = None
                self.limpiar_inputs_producto(); self.refrescar_productos()
                self.refrescar_completer()
                return

            if existe:
                # Producto existe pero no estamos en modo edición: entrar en modo edición
                self._editing_product_id = existe.id
                self.input_nombre.setText(existe.nombre)
                self.input_precio.setText(f'{existe.precio:.2f}')
                self.input_categoria.setText(existe.categoria or '')
                self.statusBar().showMessage(
                    f'Producto encontrado. Modifique los campos y presione Agregar/Actualizar.', 5000)
                self.input_nombre.setFocus()
                self.input_nombre.selectAll()
                return

            nuevo = Producto(codigo_barra=c,nombre=n,precio=pr,categoria=ca)
            self.session.add(nuevo);
            self.session.commit()
            self._sync_push("producto", nuevo)
            self._log_product_change(nuevo, 'producto', '', f'{c} / {n} / ${pr}', 'Formulario - Nuevo producto')

            datos = {'codigo_barra':c,'nombre':n,'precio':pr,'categoria':ca}
            self.history.append(('add',datos))
            self.statusBar().showMessage('Producto creado',3000)
            self._beep_ok()
            self._editing_product_id = None
            self.limpiar_inputs_producto(); self.refrescar_productos()
            self.refrescar_completer()
    def eliminar_productos(self):
        if QMessageBox.question(self,'Confirmar','¿Eliminar productos seleccionados?',
                                QMessageBox.Yes|QMessageBox.No) != QMessageBox.Yes:
            return

        deleted = []
        for r in range(self.table_productos.rowCount()):
            if self._is_row_checked(r, self.table_productos):
                pid = int(self.table_productos.item(r, 1).text())
                p   = self.session.query(Producto).get(pid)
                if p:
                    deleted.append({ 
                    'codigo_barra': p.codigo_barra,
                    'nombre': p.nombre,
                    'precio': p.precio,
                    'categoria': p.categoria
                    })
                    self.session.delete(p)

        if deleted:
            self.session.commit()
            for d in deleted:
                self._sync_push("producto_del", d["codigo_barra"])
            self.history.append(('del', deleted))
            self.statusBar().showMessage(f'{len(deleted)} productos eliminados', 3000)
            self.refrescar_productos()
            self.refrescar_completer()
            
            
    def imprimir_codigos(self):
        # Filas marcadas con el checkbox de la columna 0
        filas = [r for r in range(self.table_productos.rowCount())
                if self._is_row_checked(r, self.table_productos)]
        if not filas:
            QMessageBox.information(self, 'Imprimir', 'No hay productos seleccionados.')
            return

        # Datos a imprimir (código + nombre)
        items = []
        for r in filas:
            code = self.table_productos.item(r, 2).text()  # 'Código'
            name = self.table_productos.item(r, 3).text()  # 'Nombre'
            items.append((code, name))

        # Diálogo de impresión
        printer = QPrinter(QPrinter.HighResolution)
        dlg = QPrintDialog(printer, self)
        if dlg.exec_() != QPrintDialog.Accepted:
            return

        painter = QPainter(printer)
        try:
            page = printer.pageRect()
            margin = 24
            label_w = page.width() - 2 * margin

            from PyQt5.QtGui import QFont, QFontMetrics

            # Fuente a 12 px como pediste
            font_code = QFont("Arial"); font_code.setPixelSize(12); font_code.setBold(False)
            fm_code   = QFontMetrics(font_code)
            font_name = QFont("Arial"); font_name.setPixelSize(12); font_name.setBold(False)
            fm_name   = QFontMetrics(font_name)

            h_bar  = 120
            y = margin  # ← INICIALIZAR AQUÍ

            for code, name in items:
                # Generar el código de barras sin texto embebido
                bc = barcode.get('code128', code, writer=ImageWriter())
                buf = BytesIO()
                bc.write(buf, options={"module_height": 18.0, "write_text": False})
                pix = QPixmap(); pix.loadFromData(buf.getvalue())

                # ¿cabe en la página actual?
                block_h = h_bar + 6 + fm_code.height() + 2 + fm_name.height() + 16
                if y + block_h > page.height() - margin:
                    printer.newPage()
                    y = margin

                # 1) Barra centrada
                scaled   = pix.scaled(label_w, h_bar, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                scaled_w = scaled.width()
                x = margin + (label_w - scaled_w) // 2
                painter.drawPixmap(x, y, scaled)
                y += scaled.height() + 6

                # 2) Número (12 px, centrado con ancho de la barra)
                painter.setFont(font_code)
                rect_code = QRect(x, y, scaled_w, fm_code.height())
                painter.drawText(rect_code, Qt.AlignHCenter | Qt.AlignVCenter, code)
                y += fm_code.height() + 2

                # 3) Nombre (12 px, elide)
                painter.setFont(font_name)
                name_short = (name or '').strip()
                elided = fm_name.elidedText(name_short, Qt.ElideRight, scaled_w)
                rect_name = QRect(x, y, scaled_w, fm_name.height())
                painter.drawText(rect_name, Qt.AlignHCenter | Qt.AlignVCenter, elided)
                y += fm_name.height() + 16
        finally:
            painter.end()
            
    def _limpiar_valor_excel(self, valor, default=''):
        """Limpia un valor leido de Excel: maneja NaN, None, y espacios."""
        if valor is None:
            return default
        try:
            import math
            if isinstance(valor, float) and math.isnan(valor):
                return default
        except Exception:
            pass
        return str(valor).strip()

    def importar_productos(self):
        """Carga productos desde un archivo Excel con previsualización antes de aplicar cambios."""
        path, _ = QFileDialog.getOpenFileName(self, 'Importar Excel', '', 'Excel Files (*.xlsx *.xls)')
        if not path:
            return
        # Mostrar progress mientras se lee el Excel (puede tardar segundos con archivos grandes)
        from app.gui.progress_helpers import busy_dialog
        try:
            with busy_dialog(self, "Importando Excel", f"Leyendo archivo:\n{path}"):
                df = pd.read_excel(path, dtype={'codigo_barra': str})
        except Exception as e:
            from app.gui.error_messages import show_error
            show_error(self, "leer el archivo Excel", e, context="import_excel_read")
            return

        # Clasificar filas: nuevos, modificados, sin cambios
        nuevos = {}  # dict por codigo_barra para deduplicar
        modificados = []
        sin_cambios = 0
        filas_omitidas = 0

        for _, row in df.iterrows():
            codigo = self._limpiar_valor_excel(row.get('codigo_barra', ''))
            nombre = self._limpiar_valor_excel(row.get('nombre', ''))
            try:
                raw_precio = row.get('precio', 0)
                import math
                if isinstance(raw_precio, float) and math.isnan(raw_precio):
                    raw_precio = 0
                precio = float(raw_precio)
            except Exception:
                precio = 0.0
            categoria = self._limpiar_valor_excel(row.get('categoria', '')) or None
            if not codigo or not nombre or codigo == 'nan' or nombre == 'nan':
                filas_omitidas += 1
                continue

            prod = self.session.query(Producto).filter_by(codigo_barra=codigo).first()
            if prod:
                # Verificar si hay diferencias
                hay_diff = (
                    prod.nombre != nombre or
                    abs((prod.precio or 0) - precio) > 0.001 or
                    (prod.categoria or '') != (categoria or '')
                )
                if hay_diff:
                    modificados.append({
                        'codigo_barra': codigo,
                        'nombre_actual': prod.nombre,
                        'nombre_nuevo': nombre,
                        'precio_actual': prod.precio or 0,
                        'precio_nuevo': precio,
                        'categoria_actual': prod.categoria,
                        'categoria_nueva': categoria,
                        'prod_obj': prod,
                    })
                else:
                    sin_cambios += 1
            else:
                # Deduplicar por codigo: si ya hay uno, queda el ultimo
                nuevos[codigo] = {
                    'codigo_barra': codigo,
                    'nombre': nombre,
                    'precio': precio,
                    'categoria': categoria,
                }

        nuevos_list = list(nuevos.values())

        # Si no hay nada que importar
        if not nuevos_list and not modificados:
            msg = f'No hay cambios para aplicar.\n{sin_cambios} producto(s) ya estaban actualizados.'
            if filas_omitidas:
                msg += f'\n{filas_omitidas} fila(s) omitidas (datos incompletos).'
            QMessageBox.information(self, 'Importación', msg)
            return

        # Mostrar preview
        from app.gui.dialogs import ImportPreviewDialog
        dlg = ImportPreviewDialog(nuevos_list, modificados, sin_cambios, self)
        if dlg.exec_() != QDialog.Accepted:
            return

        modo, mod_seleccionados = dlg.resultado()
        if not modo:
            return

        try:
            cont_nuevos = 0
            cont_actualizados = 0
            sync_data = []  # guardar datos planos (no ORM) para sync en threads

            # Insertar nuevos (siempre, en ambos modos)
            for p in nuevos_list:
                prod = Producto(
                    codigo_barra=p['codigo_barra'], nombre=p['nombre'],
                    precio=p['precio'], categoria=p['categoria']
                )
                self.session.add(prod)
                cont_nuevos += 1

            # Actualizar modificados (solo si modo != 'solo_nuevos')
            if modo == 'todos' and mod_seleccionados:
                for m in mod_seleccionados:
                    prod = m['prod_obj']
                    # Log antes de modificar
                    for campo, actual, nuevo_val in [
                        ('nombre', prod.nombre, m['nombre_nuevo']),
                        ('precio', prod.precio, m['precio_nuevo']),
                        ('categoria', prod.categoria or '', m['categoria_nueva'] or ''),
                    ]:
                        if str(actual or '') != str(nuevo_val or ''):
                            self._log_product_change(prod, campo, actual, nuevo_val, 'Importar Excel - Actualizar')
                    prod.nombre = m['nombre_nuevo']
                    prod.precio = m['precio_nuevo']
                    prod.categoria = m['categoria_nueva']
                    cont_actualizados += 1

            self.session.commit()

            # Log de nuevos y recoger datos para sync (después del commit para tener IDs)
            for p in nuevos_list:
                prod = self.session.query(Producto).filter_by(codigo_barra=p['codigo_barra']).first()
                if prod:
                    self._log_product_change(
                        prod, 'producto', '',
                        f"{p['codigo_barra']} / {p['nombre']} / ${p['precio']}",
                        'Importar Excel - Nuevo producto'
                    )
                    sync_data.append(prod)

            # Sync de modificados
            if modo == 'todos' and mod_seleccionados:
                for m in mod_seleccionados:
                    sync_data.append(m['prod_obj'])

            # Sync batch: 1 solo PATCH en vez de N requests seriados (mucho más rápido)
            if sync_data:
                try:
                    self._sync_push_batch("producto", sync_data)
                except Exception:
                    # Fallback al loop original si algo falla
                    for prod in sync_data:
                        try:
                            self._sync_push("producto", prod)
                        except Exception:
                            pass

            self.refrescar_productos()
            msg = (f'Productos nuevos: {cont_nuevos}\n'
                   f'Productos actualizados: {cont_actualizados}\n'
                   f'Sin cambios: {sin_cambios}')
            if filas_omitidas:
                msg += f'\nFilas omitidas: {filas_omitidas}'
            QMessageBox.information(self, 'Importación completada', msg)
            self.refrescar_completer()

        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(
                self, 'Error en importación',
                f'Ocurrió un error durante la importación:\n{e}\n\n'
                f'Los cambios fueron revertidos.'
            )
            self.refrescar_productos()
        
    def _autofit_openpyxl(ws, df, min_width=8, max_width=50, padding=2):
        """
        Ajusta el ancho de columnas de una hoja openpyxl en base a los contenidos de df.
        ws: hoja openpyxl (workbook.active o writer.sheets['Productos'])
        df: pandas.DataFrame exportado
        """
        from openpyxl.utils import get_column_letter

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name))
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, values_only=True):
                for val in cell:
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
            width = max(min_width, min(max_len + padding, max_width))
            ws.column_dimensions[get_column_letter(col_idx)].width = width


    def exportar_productos(self):
        """Exporta productos a Excel. Pregunta si exportar todos, seleccionados o filtrados."""
        from PyQt5.QtWidgets import QMessageBox, QFileDialog
        import pandas as pd, os

        # Determinar si hay seleccionados o filtro activo
        n_sel = len(self._selected_product_ids)
        texto_filtro = self.input_buscar.text().strip()
        hay_filtro = bool(texto_filtro)

        # Preguntar qué exportar si hay seleccionados o filtro
        modo = 'todos'
        if n_sel > 0 or hay_filtro:
            opciones = ['Todos los productos']
            if n_sel > 0:
                opciones.append(f'Solo seleccionados ({n_sel})')
            if hay_filtro:
                n_filtrados = self.table_productos.rowCount()
                opciones.append(f'Solo filtrados ({n_filtrados}) - "{texto_filtro}"')
            opciones.append('Cancelar')

            from PyQt5.QtWidgets import QInputDialog
            opcion, ok = QInputDialog.getItem(
                self, 'Exportar a Excel', '¿Qué productos exportar?',
                opciones, 0, False
            )
            if not ok or opcion == 'Cancelar':
                return
            if opcion.startswith('Solo seleccionados'):
                modo = 'seleccionados'
            elif opcion.startswith('Solo filtrados'):
                modo = 'filtrados'

        path, _ = QFileDialog.getSaveFileName(self, 'Exportar Excel', 'productos.xlsx', 'Excel Files (*.xlsx)')
        if not path:
            return

        # Obtener productos según el modo
        if modo == 'seleccionados':
            productos = [
                self.session.query(Producto).get(pid)
                for pid in self._selected_product_ids
            ]
            productos = [p for p in productos if p]
        elif modo == 'filtrados':
            # Leer IDs de la tabla visible
            ids = []
            for r in range(self.table_productos.rowCount()):
                it = self.table_productos.item(r, 1)
                if it:
                    try:
                        ids.append(int(it.text()))
                    except Exception:
                        pass
            productos = [
                self.session.query(Producto).get(pid)
                for pid in ids
            ]
            productos = [p for p in productos if p]
        else:
            productos = self.session.query(Producto).all()

        data = [{
            'codigo_barra': p.codigo_barra,
            'nombre':       p.nombre,
            'precio':       p.precio,
            'categoria':    p.categoria or ''
        } for p in productos]
        df = pd.DataFrame(data)

        # Guardado con openpyxl y auto-fit (si está disponible)
        try:
            from openpyxl import load_workbook
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Productos')
                ws = writer.sheets['Productos']
                try:
                    self._autofit_openpyxl(ws, df, min_width=8, max_width=50, padding=2)
                except Exception:
                    pass
            QMessageBox.information(
                self, 'Exportar Excel',
                f'Guardado: {path}\n{len(productos)} producto(s) exportados.'
            )
        except Exception as e:
            # Fallback a CSV si openpyxl no está disponible
            try:
                base, _ = os.path.splitext(path)
                csv_path = base + '.csv'
                df.to_csv(csv_path, index=False, encoding='utf-8')
                QMessageBox.information(
                    self, 'Exportar CSV',
                    f'No se pudo escribir Excel ({e}).\nGuardado CSV: {csv_path}\n'
                    f'{len(productos)} producto(s) exportados.'
                )
            except Exception as e2:
                QMessageBox.warning(self, 'Exportar', f'No se pudo guardar:\n{e2}')

        
    def _limpiar_campos_productos(self):
        """Limpia todos los campos de búsqueda y formulario de productos (tecla ESC)."""
        try:
            self.input_buscar.clear()
            self.input_codigo.clear()
            self.input_nombre.clear()
            self.input_precio.clear()
            self.input_categoria.clear()
            # Resetear modo edición si existe
            if hasattr(self, '_editing_product_id'):
                self._editing_product_id = None
            # Limpiar selección persistente
            self._selected_product_ids.clear()
            # Refrescar tabla SIN preservar selección (para que no re-lea los checks)
            self.refrescar_productos(preserve_selection=False)
            # Foco al buscador
            self.input_buscar.setFocus()
        except Exception:
            pass

    def _update_completer_productos(self):
        """Actualiza el filtro del proxy del completer de productos tras debounce."""
        try:
            text = self.input_buscar.text().strip()
            if len(text) < 2:
                return
            comp = getattr(self, '_completer_productos', None)
            if not comp:
                return
            proxy = getattr(comp, '_proxy', None)
            if proxy:
                proxy.setFilterWildcard(f"*{text}*")
                comp.setCompletionPrefix(text)
                comp.complete()
        except Exception:
            pass

    def _on_buscar_text_changed(self, txt):
        """Reinicia el debounce timer cada vez que cambia el texto."""
        self._search_timer.start()

    def _do_buscar_productos(self):
        """Ejecuta la busqueda tras el debounce."""
        self.buscar_productos(self.input_buscar.text())

    def buscar_productos(self, txt):
            self.refrescar_productos()

            # Si hay búsqueda pero 0 resultados, ofrecer agregar
            if txt and txt.strip() and self.table_productos.rowCount() == 0:
                search_term = txt.strip()
                if " - " in search_term:
                    search_term = search_term.split(" - ")[0].strip()
                from PyQt5.QtWidgets import QMessageBox
                resp = QMessageBox.question(
                    self, "Producto no encontrado",
                    f'No se encontro "{search_term}".\n¿Desea agregarlo?',
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                if resp == QMessageBox.Yes:
                    from app.gui.dialogs import agregar_producto_rapido_dialog
                    nuevo = agregar_producto_rapido_dialog(
                        self.session, self, term=search_term,
                        sync_push_fn=self._sync_push,
                        completer_refresh_fn=getattr(self, 'refrescar_completer', None),
                    )
                    if nuevo:
                        # Log solo si es producto recien creado (no uno existente reutilizado)
                        if not getattr(nuevo, '_from_existing', False):
                            self._log_product_change(
                                nuevo, 'producto', '',
                                f'{nuevo.codigo_barra} / {nuevo.nombre} / ${nuevo.precio}',
                                'Dialogo rapido - Nuevo producto'
                            )
                        self.input_buscar.clear()
                        self.refrescar_productos()
                        self.statusBar().showMessage(f'Producto "{nuevo.nombre}" creado', 3000)

    def cargar_producto(self,row,col):
        tbl = self.table_productos
        
            # 1) Intentar leer el ID de la columna 1 (si está visible)
        pid = None
        try:
            it_id = tbl.item(row, 1)
            if it_id:
                txt = (it_id.text() or "").strip()
                if txt.isdigit():
                    pid = int(txt)
        except Exception:
            pid = None

        # 2) Si no hay ID visible, recuperar el ID guardado en UserRole de Código/Nombre
        if pid is None:
            # primero columna Código (2), si no, Nombre (3)
            for c in (2, 3):
                it = tbl.item(row, c)
                if it is not None:
                    pid = it.data(Qt.UserRole)
                    if pid:
                        break

        if not pid:
            return  # no podemos cargar

        # 3) Traer el producto y poblar el formulario
        try:
            p = self.session.query(Producto).get(pid)  # si usás SQLAlchemy 2.x, usa Session.get(Producto, pid)
        except Exception:
            p = None

        if not p:
            return

        self._editing_product_id = pid
        self.input_codigo.setText(p.codigo_barra or '')
        self.input_nombre.setText(p.nombre or '')
        self.input_precio.setText(f'{(p.precio or 0.0):.2f}')
        self.input_categoria.setText(p.categoria or '')
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(0, lambda: self.table_productos.clearSelection())

    def limpiar_inputs_producto(self):
        self._editing_product_id = None
        for fld in (self.input_codigo,self.input_nombre,
                    self.input_precio,self.input_categoria):
            fld.clear()
        self.input_codigo.setFocus()
            
    def _find_row_by_codigo(self, codigo: str):
        for r in range(self.table_cesta.rowCount()):
            it = self.table_cesta.item(r, 0)
            if it and it.text() == codigo:
                return r
        return None
    
    
        
    def _on_producto_item_changed(self, item):
        row, col = item.row(), item.column()
        if col not in (3, 4, 5):
            return
        tbl = self.table_productos

        try:
            prod_id = int(tbl.item(row, 1).text())
        except Exception:
            return

        from PyQt5.QtCore import QSignalBlocker
        blocker = QSignalBlocker(tbl)
        try:
            prod = self.prod_repo.obtener(prod_id)
            if not prod:
                return

            campo = {3: 'nombre', 4: 'precio', 5: 'categoria'}[col]
            anterior = getattr(prod, campo)

            if col == 4:
                try:
                    nuevo_precio = float(item.text().replace(',', '.'))
                except Exception:
                    item.setText(f"{prod.precio:.2f}")
                    return
                self.prod_repo.actualizar_precio(prod_id, nuevo_precio)
                item.setText(f"{nuevo_precio:.2f}")
                nuevo = nuevo_precio
            elif col == 3:
                nuevo = item.text().strip()
                self.prod_repo.actualizar_nombre(prod_id, nuevo)
            elif col == 5:
                nuevo = item.text().strip()
                self.prod_repo.actualizar_categoria(prod_id, nuevo)

            # Log del cambio individual
            self._log_product_change(prod, campo, anterior, nuevo, f'Edición directa - {campo.capitalize()}')

            # Sync: publicar producto editado
            prod = self.prod_repo.obtener(prod_id)
            if prod:
                self._sync_push("producto", prod)
        finally:
            del blocker
            
            
    def refrescar_productos(self, preserve_selection=True):
        from app.utils_timing import measure
        with measure("refrescar_productos"):
            # 1) Filtro — usar SQL LIKE para no cargar 13K objetos
            texto_busqueda = self.input_buscar.text().strip()
            if " - " in texto_busqueda:
                texto_busqueda = texto_busqueda.split(" - ")[0].strip()
            texto_busqueda = texto_busqueda.lower()
            self.productos_filtro = texto_busqueda

            if texto_busqueda:
                # Buscar con SQL LIKE (indice en la DB, mucho mas rapido)
                productos = self.prod_repo.buscar(texto_busqueda)
            else:
                productos = self.prod_repo.listar_todos()

            total = len(productos)

            # 2) Pintar tabla: 0=Sel, 1=ID, 2=Código, 3=Nombre, 4=Precio, 5=Categoría
            tbl = self.table_productos

            tbl.setColumnCount(6)
            tbl.setHorizontalHeaderLabels(
                ['Sel','ID','Código','Nombre','Precio','Categoría']
            )

            # Preservar selección antes de rebuild (salvo si ESC la limpió)
            if preserve_selection:
                for r in range(tbl.rowCount()):
                    id_item = tbl.item(r, 1)
                    if id_item:
                        try:
                            pid = int(id_item.text())
                            if self._is_row_checked(r, tbl):
                                self._selected_product_ids.add(pid)
                            else:
                                self._selected_product_ids.discard(pid)
                        except (ValueError, AttributeError):
                            pass

            from app.gui.qt_helpers import freeze_table
            with freeze_table(tbl):
                tbl.setRowCount(0)
                tbl.setRowCount(total)
                for r, p in enumerate(productos):
                    # Col 0: checkbox (restaurar selección si estaba marcado)
                    chk = QTableWidgetItem()
                    chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                    chk.setCheckState(Qt.Checked if p.id in self._selected_product_ids else Qt.Unchecked)
                    tbl.setItem(r, 0, chk)

                    # Cols 1..5: ID, Código, Nombre, Precio, Categoría
                    vals = [p.id, p.codigo_barra, p.nombre, f'{p.precio:.2f}', p.categoria or '']
                    for c, val in enumerate(vals, start=1):
                        it = QTableWidgetItem(str(val))
                        it.setFont(QFont("Arial", 11))
                        if c in (3, 4, 5):  # Nombre, Precio, Categoría editables
                            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                        else:
                            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                        tbl.setItem(r, c, it)

            # 3) Footer con contador de seleccionados
            n_sel = len(self._selected_product_ids)
            extra = f" | {n_sel} seleccionados" if n_sel > 0 else ""
            self.lbl_paginacion.setText(f"Total: {total} productos{extra}")
            
    def _abrir_ultimos_cambios(self):
        """Abre diálogo con el historial de ediciones masivas de la sesión."""
        from app.gui.dialogs import UltimosCambiosDialog
        log = getattr(self, '_product_change_log', [])
        dlg = UltimosCambiosDialog(log, self)
        dlg.exec_()

    def _productos_editar_precio_buscado(self):
        """Atajo Ñ (Productos): edita el producto seleccionado/buscado usando QuickEditProductoDialog.

        Prioridad:
          (1) primera fila seleccionada en la tabla de productos;
          (2) texto del buscador `self.input_buscar`.
        Guarda los cambios, loguea, sincroniza y refresca la UI.
        """
        from app.gui.dialogs import QuickEditProductoDialog
        prod = None

        # (1) Producto seleccionado en la tabla (columna 2 = Código)
        try:
            sel = self.table_productos.selectionModel().selectedRows()
            if sel:
                row = sel[0].row()
                it_code = self.table_productos.item(row, 2)
                if it_code and it_code.text():
                    prod = self.prod_repo.buscar_por_codigo(it_code.text().strip())
        except Exception:
            prod = None

        # (2) Fallback: texto del buscador
        if not prod:
            try:
                term = (self.input_buscar.text() or "").strip()
            except Exception:
                term = ""
            if not term:
                try:
                    self.statusBar().showMessage(
                        'Seleccioná un producto o escribí código/nombre en el buscador.', 2500)
                except Exception:
                    pass
                return
            # Lookup: exacto por código si es numérico; búsqueda por nombre en otro caso
            if term.isdigit():
                prod = self.prod_repo.buscar_por_codigo(term)
            else:
                try:
                    resultados = self.prod_repo.buscar(term, limit=1)
                    prod = resultados[0] if resultados else None
                except Exception:
                    prod = None
            if not prod:
                QMessageBox.information(self, 'No encontrado', f'Producto "{term}" no hallado.')
                return

        viejos = {
            'nombre': prod.nombre,
            'precio': prod.precio,
            'categoria': prod.categoria,
        }
        dlg = QuickEditProductoDialog(prod, self)
        dlg.setWindowTitle(f'Editar precio - {prod.nombre}')
        if dlg.exec_() != QDialog.Accepted:
            return
        tup = dlg.datos() or (prod.nombre, prod.precio, prod.categoria)
        nom, precio_nuevo, cat = tup[0], tup[1], tup[2]
        try:
            prod.nombre = nom
            prod.precio = precio_nuevo
            prod.categoria = cat
            if hasattr(prod, 'version'):
                prod.version = (prod.version or 1) + 1
            self.session.commit()
        except Exception as e:
            QMessageBox.warning(self, 'Error', f'No se pudo guardar: {e}')
            return
        # Log de cambios
        try:
            for k, v_old, v_new in [
                ('nombre', viejos['nombre'], nom),
                ('precio', viejos['precio'], precio_nuevo),
                ('categoria', viejos['categoria'] or '', cat or ''),
            ]:
                if str(v_old or '') != str(v_new or ''):
                    self._log_product_change(prod, k, v_old, v_new, 'Productos - Atajo Ñ')
        except Exception:
            pass
        # Sync
        try:
            self._sync_push("producto", prod)
        except Exception:
            pass
        # Refresco UI
        try:
            self.refrescar_productos()
        except Exception:
            pass
        try:
            self.refrescar_completer()
        except Exception:
            pass
        try:
            self.statusBar().showMessage(
                f'Precio de {prod.nombre} actualizado a ${float(precio_nuevo):.2f}', 3000)
        except Exception:
            pass

    def _log_product_change(self, prod, campo, anterior, nuevo, operacion):
        """Registra un cambio individual en el log de la sesión."""
        import datetime as _dt
        if not hasattr(self, '_product_change_log'):
            self._product_change_log = []
        ant_str = f"{anterior:.2f}" if isinstance(anterior, float) else str(anterior or '')
        nue_str = f"{nuevo:.2f}" if isinstance(nuevo, float) else str(nuevo or '')
        if ant_str == nue_str:
            return
        self._product_change_log.append({
            'fecha': _dt.datetime.now(),
            'operacion': operacion,
            'cambios': [{
                'producto_id': prod.id,
                'codigo_barra': prod.codigo_barra,
                'nombre': prod.nombre,
                'campo': campo,
                'anterior': ant_str,
                'nuevo': nue_str,
            }]
        })

    def _ver_productos_sin_ventas(self):
        """Muestra productos que no se vendieron en los últimos 90 días."""
        from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem,
                                      QHeaderView, QDialogButtonBox, QLabel, QSpinBox, QHBoxLayout)
        from app.models import VentaItem, Venta
        from sqlalchemy import func
        from datetime import datetime, timedelta

        dlg = QDialog(self)
        dlg.setWindowTitle("Productos sin movimiento")
        dlg.resize(800, 500)
        layout = QVBoxLayout(dlg)

        # Control de días
        row = QHBoxLayout()
        row.addWidget(QLabel("Sin ventas en los últimos"))
        spin_dias = QSpinBox()
        spin_dias.setRange(7, 365)
        spin_dias.setValue(90)
        spin_dias.setSuffix(" días")
        row.addWidget(spin_dias)
        btn_buscar = QPushButton("Buscar")
        row.addWidget(btn_buscar)
        row.addStretch()
        layout.addLayout(row)

        lbl_info = QLabel("")
        layout.addWidget(lbl_info)

        table = QTableWidget(0, 5)
        table.setHorizontalHeaderLabels(["Código", "Nombre", "Precio", "Categoría", "Última venta"])
        table.verticalHeader().setVisible(False)
        hdr = table.horizontalHeader()
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)
        for c in [0, 2, 3, 4]:
            hdr.setSectionResizeMode(c, QHeaderView.ResizeToContents)
        layout.addWidget(table)

        def _buscar():
            dias = spin_dias.value()
            limite = datetime.now() - timedelta(days=dias)

            # Subquery: productos que SÍ tuvieron ventas
            vendidos = (
                self.session.query(VentaItem.producto_id)
                .join(Venta, Venta.id == VentaItem.venta_id)
                .filter(Venta.fecha >= limite)
                .filter(VentaItem.producto_id.isnot(None))
                .distinct()
                .subquery()
            )

            from app.models import Producto
            sin_ventas = (
                self.session.query(Producto)
                .filter(~Producto.id.in_(self.session.query(vendidos.c.producto_id)))
                .order_by(Producto.nombre)
                .all()
            )

            table.setRowCount(0)
            for i, p in enumerate(sin_ventas):
                table.insertRow(i)
                table.setItem(i, 0, QTableWidgetItem(p.codigo_barra))
                table.setItem(i, 1, QTableWidgetItem(p.nombre))
                table.setItem(i, 2, QTableWidgetItem(f"${p.precio:.2f}"))
                table.setItem(i, 3, QTableWidgetItem(p.categoria or ""))

                # Buscar última venta de este producto (si existe)
                ultima = (
                    self.session.query(func.max(Venta.fecha))
                    .join(VentaItem, VentaItem.venta_id == Venta.id)
                    .filter(VentaItem.producto_id == p.id)
                    .scalar()
                )
                if ultima:
                    table.setItem(i, 4, QTableWidgetItem(ultima.strftime("%d/%m/%Y")))
                else:
                    table.setItem(i, 4, QTableWidgetItem("Nunca"))

            lbl_info.setText(f"Se encontraron {len(sin_ventas)} productos sin ventas en los últimos {dias} días.")

        btn_buscar.clicked.connect(_buscar)
        _buscar()  # Ejecutar búsqueda inicial

        btns = QDialogButtonBox(QDialogButtonBox.Close, dlg)
        btns.rejected.connect(dlg.close)
        layout.addWidget(btns)

        dlg.exec_()

