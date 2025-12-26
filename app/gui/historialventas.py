# app/gui/historialventas.py
# -*- coding: utf-8 -*-
from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime, timedelta, time as dtime
from typing import List, Optional


import os
import tempfile

import pandas as pd
from PyQt5.QtCore import Qt, QTimer, QDate,QTime

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QLabel, QLineEdit, QPushButton,
    QComboBox, QDateEdit, QTableWidget, QTableWidgetItem, QHeaderView, QCheckBox,
    QFileDialog, QMessageBox, QDialog, QTableWidgetSelectionRange,QTimeEdit, QSpinBox,
    QTabWidget, QScrollArea, QFrame, QGroupBox, QGridLayout
)

from app.config import load as load_config, save as save_config   # ← config existente :contentReference[oaicite:1]{index=1}
from app.models import Venta, VentaItem
from app.repository import VentaRepo                               # ← repo existente (listar_por_rango, listar_items) :contentReference[oaicite:2]{index=2}



# === Helpers Excel (definidos a nivel de módulo) ===
def _make_writer(path):
    import pandas as pd
    # El caller ya asegura .xlsx, acá solo probamos engines
    for engine in ("openpyxl", "xlsxwriter"):
        try:
            return pd.ExcelWriter(path, engine=engine)
        except Exception:
            continue
    return None


def _autofit_sheet(ws, df, engine_name: str):
    def _maxlen(series):
        try:
            return max([len(str(series.name))] + [len(str(x)) for x in series.astype(str).tolist()])
        except Exception:
            return 12

    if "openpyxl" in (engine_name or "").lower():
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(df.columns, start=1):
            width = min(max(10, _maxlen(df[col]) + 2), 60)
            ws.column_dimensions[get_column_letter(i)].width = width
    else:  # xlsxwriter
        for i, col in enumerate(df.columns):
            width = min(max(10, _maxlen(df[col]) + 2), 60)
            try:
                ws.set_column(i, i, width)
            except Exception:
                pass

def _get_any(obj, names, default=None):
    for n in names:
        try:
            v = getattr(obj, n, None)
        except Exception:
            v = None
        if v not in (None, ""):
            return v
    return default

def _norm_item_fields(it):
    # Códigos posibles en VentaItem
    codigo = _get_any(it, ['codigo', 'codigo_barra', 'cod_barra', 'codigobarra', 'cod'], "") or ""
    # Nombre directo o a través de la relación producto
    nombre = _get_any(it, ['nombre'], "")
    if not nombre:
        prod = getattr(it, 'producto', None)
        if prod:
            nombre = _get_any(prod, ['nombre', 'descripcion', 'desc'], "") or ""
    # Cantidad / precio unitario con alias típicos
    try:
        cant = float(_get_any(it, ['cantidad', 'cant', 'unidades'], 1) or 1)
    except Exception:
        cant = 1.0
    try:
        pu = float(_get_any(it, ['precio', 'precio_unit', 'precio_unitario', 'p_unit', 'precioU'], 0.0) or 0.0)
    except Exception:
        pu = 0.0
    return codigo, nombre, cant, pu
# ---------------------- Email helper (SMTP simple) ----------------------
def _send_mail_with_attachments(subject: str, body: str,
                                recipients: List[str], attachments: List[str] = None):
    import smtplib
    from email.message import EmailMessage

    cfg = load_config()
    e = (cfg.get("email") or {})
    smtp = (e.get("smtp") or {})

    host = smtp.get("host")
    port = int(smtp.get("port") or 587)
    use_tls = bool(smtp.get("use_tls", True))
    user = smtp.get("username")
    pwd  = smtp.get("password")
    from_addr = e.get("sender") or user

    if not host or not user or not pwd or not from_addr:
        raise RuntimeError("Falta configurar SMTP en Configuración → Email.")

    msg = EmailMessage()
    msg["Subject"] = subject or "Reporte"
    msg["From"] = from_addr
    msg["To"] = ", ".join(recipients or [])
    if e.get("bcc"):
        msg["Bcc"] = ", ".join(e.get("bcc"))

    msg.set_content(body or "")

    for path in (attachments or []):
        try:
            with open(path, "rb") as f:
                data = f.read()
            filename = os.path.basename(path)
            msg.add_attachment(data, maintype="application",
                               subtype="octet-stream", filename=filename)
        except Exception as ex:
            print("Adjunto falló:", ex)

    if use_tls:
        server = smtplib.SMTP(host, port)
        server.starttls()
    else:
        server = smtplib.SMTP_SSL(host, port)

    server.login(user, pwd)
    server.send_message(msg)
    server.quit()
    return True


# ---------------------- UI principal ----------------------
class HistorialVentasWidget(QWidget):
    """
    Pestaña Historial de ventas:
    - Filtros por fecha (desde/hasta), sucursal, forma pago, texto.
    - Tabla con ventas del rango.
    - Botón Enviar a correo (Excel con filtros aplicados).
    - Programación de envíos (diario/semanal/mensual a hora fija).
    - Doble clic en una venta => dialog con items vendidos.
    """
    def __init__(self, session, sucursal_actual: Optional[str] = None, parent=None, es_admin: bool = False):
        super().__init__(parent)
        self.session = session
        self.repo = VentaRepo(self.session)
        self.sucursal_actual = sucursal_actual
        self.es_admin = es_admin

        self._ventas_cache: List[Venta] = []

        root = QVBoxLayout(self)

        # Crear QTabWidget para Listado y Estadísticas
        self.tabs = QTabWidget()
        root.addWidget(self.tabs)

        # TAB 1: LISTADO
        tab_listado = QWidget()
        lay_listado = QVBoxLayout(tab_listado)

        # --- Filtros ---
        row1 = QHBoxLayout()
        self.dt_desde = QDateEdit()
        self.dt_hasta = QDateEdit()
        for d in (self.dt_desde, self.dt_hasta):
            d.setDisplayFormat("yyyy-MM-dd")
            d.setCalendarPopup(True)
        hoy = QDate.currentDate()
        self.dt_desde.setDate(hoy)
        self.dt_hasta.setDate(hoy)

        self.cmb_sucursal = QComboBox()
        self.cmb_sucursal.addItem("Todas", None)
        for s in ("Sarmiento", "Salta"):
            self.cmb_sucursal.addItem(s, s)
        if self.sucursal_actual and self.sucursal_actual in ("Sarmiento", "Salta"):
            self.cmb_sucursal.setCurrentText(self.sucursal_actual)

        self.cmb_forma = QComboBox()
        self.cmb_forma.addItems(["Todas", "Efectivo", "Tarjeta"])

        self.txt_buscar = QLineEdit()
        self.txt_buscar.setPlaceholderText("Buscar por Nº ticket o texto...")

        row1.addWidget(QLabel("Desde:")); row1.addWidget(self.dt_desde)
        row1.addWidget(QLabel("Hasta:")); row1.addWidget(self.dt_hasta)
        row1.addWidget(QLabel("Sucursal:")); row1.addWidget(self.cmb_sucursal)
        row1.addWidget(QLabel("Forma:")); row1.addWidget(self.cmb_forma)
        row1.addWidget(self.txt_buscar)

        btn_filtrar = QPushButton("Aplicar filtros")
        btn_filtrar.clicked.connect(self.refrescar)
        row1.addWidget(btn_filtrar)
        lay_listado.addLayout(row1)

        # Botones de rango rápido
        row_quick = QHBoxLayout()
        row_quick.addWidget(QLabel("Rango rápido:"))

        btn_hoy = QPushButton("Hoy")
        btn_hoy.clicked.connect(self._set_rango_hoy)
        row_quick.addWidget(btn_hoy)

        btn_semana = QPushButton("Esta Semana")
        btn_semana.clicked.connect(self._set_rango_semana)
        row_quick.addWidget(btn_semana)

        btn_mes = QPushButton("Este Mes")
        btn_mes.clicked.connect(self._set_rango_mes)
        row_quick.addWidget(btn_mes)

        btn_mes_anterior = QPushButton("Mes Anterior")
        btn_mes_anterior.clicked.connect(self._set_rango_mes_anterior)
        row_quick.addWidget(btn_mes_anterior)

        row_quick.addStretch()
        lay_listado.addLayout(row_quick)

        # refrescar con Enter en el texto
        self.txt_buscar.returnPressed.connect(self.refrescar)
        # refrescar al cambiar fechas/combos
        self.dt_desde.dateChanged.connect(lambda *_: self.refrescar())
        self.dt_hasta.dateChanged.connect(lambda *_: self.refrescar())
        self.cmb_sucursal.currentIndexChanged.connect(lambda *_: self.refrescar())
        self.cmb_forma.currentIndexChanged.connect(lambda *_: self.refrescar())

        # --- Tabla ---
        self.tbl = QTableWidget(0, 15)
        self.tbl.setHorizontalHeaderLabels([
            "Nº Ticket", "Fecha/Hora", "Sucursal", "Forma Pago",
            "Cuotas","Interés", "Descuento", "Monto x cuota", "Total", "Pagado", "Vuelto", "CAE", "Vto CAE", "Comentario", "ID"
        ])
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Stretch)
        for col in (5, 6, 7):  # Interés, Descuento, Monto x cuota
            hdr.setSectionResizeMode(col, QHeaderView.ResizeToContents)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)
        self.tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tbl.cellDoubleClicked.connect(self._ver_detalle_venta)
        lay_listado.addWidget(self.tbl)

        # --- Barra inferior ---
        bar = QHBoxLayout()
        self.lbl_resumen = QLabel("0 ventas — Total $0.00")
        bar.addWidget(self.lbl_resumen)
        bar.addStretch(1)

        self.chk_incluir_items = QCheckBox("Incluir detalle de productos en Excel")
        bar.addWidget(self.chk_incluir_items)

        self.btn_excel = QPushButton("Enviar a correo")
        self.btn_excel.clicked.connect(self._exportar_y_enviar)
        bar.addWidget(self.btn_excel)

        self.btn_guardar_xlsx = QPushButton("Exportar XLSX…")
        self.btn_guardar_xlsx.clicked.connect(self._exportar_a_xlsx_local)
        bar.addWidget(self.btn_guardar_xlsx)

        lay_listado.addLayout(bar)

        # Agregar tab de listado
        self.tabs.addTab(tab_listado, "Listado")

        # TAB 2: ESTADÍSTICAS (solo para administradores)
        if self.es_admin:
            try:
                tab_stats = self._crear_tab_estadisticas()
                self.tabs.addTab(tab_stats, "Estadísticas")
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creando tab de estadísticas: {e}", exc_info=True)
                # Crear un tab de error simple
                error_widget = QWidget()
                error_layout = QVBoxLayout(error_widget)
                error_label = QLabel(f"Error al cargar estadísticas: {str(e)}")
                error_label.setStyleSheet("color: red; padding: 20px;")
                error_layout.addWidget(error_label)
                self.tabs.addTab(error_widget, "Estadísticas (Error)")

            # Conectar evento de cambio de pestaña para actualizar estadísticas automáticamente
            self.tabs.currentChanged.connect(self._on_tab_changed)

        self.refrescar()

    def _on_tab_changed(self, index):
        """Se ejecuta cuando cambia la pestaña activa"""
        # Si cambia a la pestaña de Estadísticas (índice 1), actualizar automáticamente
        if index == 1:
            try:
                self._actualizar_estadisticas()
            except Exception as e:
                print(f"Error al actualizar estadísticas automáticamente: {e}")

    # Métodos para botones de rango rápido
    def _set_rango_hoy(self):
        """Establece el rango de fechas a hoy"""
        hoy = QDate.currentDate()
        self.dt_desde.setDate(hoy)
        self.dt_hasta.setDate(hoy)

    def _set_rango_semana(self):
        """Establece el rango de fechas a esta semana (lunes a domingo)"""
        hoy = QDate.currentDate()
        # Calcular el lunes de esta semana
        dias_desde_lunes = hoy.dayOfWeek() - 1  # Qt: lunes=1, domingo=7
        lunes = hoy.addDays(-dias_desde_lunes)
        domingo = lunes.addDays(6)
        self.dt_desde.setDate(lunes)
        self.dt_hasta.setDate(domingo)

    def _set_rango_mes(self):
        """Establece el rango de fechas a este mes"""
        hoy = QDate.currentDate()
        primer_dia = QDate(hoy.year(), hoy.month(), 1)
        # Último día del mes
        if hoy.month() == 12:
            ultimo_dia = QDate(hoy.year(), 12, 31)
        else:
            ultimo_dia = QDate(hoy.year(), hoy.month() + 1, 1).addDays(-1)
        self.dt_desde.setDate(primer_dia)
        self.dt_hasta.setDate(ultimo_dia)

    def _set_rango_mes_anterior(self):
        """Establece el rango de fechas al mes anterior"""
        hoy = QDate.currentDate()
        if hoy.month() == 1:
            mes_anterior = 12
            año_anterior = hoy.year() - 1
        else:
            mes_anterior = hoy.month() - 1
            año_anterior = hoy.year()

        primer_dia = QDate(año_anterior, mes_anterior, 1)
        # Último día del mes anterior
        if mes_anterior == 12:
            ultimo_dia = QDate(año_anterior, 12, 31)
        else:
            ultimo_dia = QDate(año_anterior, mes_anterior + 1, 1).addDays(-1)
        self.dt_desde.setDate(primer_dia)
        self.dt_hasta.setDate(ultimo_dia)

    def _crear_tab_estadisticas(self):
        """Crea el tab de estadísticas con gráficos y KPIs"""
        container = QWidget()
        scroll = QScrollArea()
        scroll.setWidget(container)
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        lay = QVBoxLayout(container)

        # Banner de filtros activos
        self.stats_filtros_banner = QLabel("")
        self.stats_filtros_banner.setStyleSheet("""
            QLabel {
                background-color: #e3f2fd;
                padding: 8px;
                border-radius: 4px;
                font-weight: bold;
                color: #1976d2;
            }
        """)
        lay.addWidget(self.stats_filtros_banner)

        # Botón para actualizar estadísticas
        btn_actualizar = QPushButton("Actualizar Estadísticas")
        btn_actualizar.clicked.connect(self._actualizar_estadisticas)
        lay.addWidget(btn_actualizar)

        # KPI Cards
        kpis_layout = QHBoxLayout()
        self.kpi_total = self._crear_kpi_card("Total Vendido", "$0.00", "#2e7d32")
        self.kpi_cantidad = self._crear_kpi_card("Cantidad de Ventas", "0", "#1976d2")
        self.kpi_promedio = self._crear_kpi_card("Promedio por Venta", "$0.00", "#f57c00")
        self.kpi_interes = self._crear_kpi_card("Total Intereses", "$0.00", "#c62828")
        kpis_layout.addWidget(self.kpi_total)
        kpis_layout.addWidget(self.kpi_cantidad)
        kpis_layout.addWidget(self.kpi_promedio)
        kpis_layout.addWidget(self.kpi_interes)
        lay.addLayout(kpis_layout)

        # Grupo para gráfico de barras
        chart_group = QGroupBox("Ventas por Día")
        chart_group_layout = QVBoxLayout(chart_group)
        self.stats_chart_container = QWidget()
        self.stats_chart_layout = QVBoxLayout(self.stats_chart_container)
        self.stats_chart_layout.setContentsMargins(0, 0, 0, 0)
        chart_group_layout.addWidget(self.stats_chart_container)
        lay.addWidget(chart_group)

        # Grupo para gráfico de torta (formas de pago)
        pie_group = QGroupBox("Distribución por Forma de Pago")
        pie_group_layout = QVBoxLayout(pie_group)
        self.stats_pie_container = QWidget()
        self.stats_pie_layout = QVBoxLayout(self.stats_pie_container)
        self.stats_pie_layout.setContentsMargins(0, 0, 0, 0)
        pie_group_layout.addWidget(self.stats_pie_container)
        lay.addWidget(pie_group)

        # Sección de comparativa (solo visible cuando se selecciona "Todas" las sucursales)
        self.stats_comparativa_group = QGroupBox("Comparativa por Sucursal")
        self.stats_comparativa_layout = QVBoxLayout(self.stats_comparativa_group)
        self.stats_comparativa_layout.setSpacing(10)
        self.stats_comparativa_layout.setContentsMargins(10, 15, 10, 15)
        self.stats_comparativa_group.setVisible(False)  # Oculto por defecto
        lay.addWidget(self.stats_comparativa_group)

        # Top 10 productos
        top_group = QGroupBox("Top 10 Productos Más Vendidos")
        top_layout = QVBoxLayout(top_group)
        top_layout.setSpacing(10)
        top_layout.setContentsMargins(10, 15, 10, 15)

        self.stats_top_productos = QTableWidget(0, 3)
        self.stats_top_productos.setHorizontalHeaderLabels(["Producto", "Cantidad", "Total"])
        self.stats_top_productos.horizontalHeader().setStretchLastSection(True)
        self.stats_top_productos.verticalHeader().setVisible(False)
        self.stats_top_productos.setSelectionMode(QTableWidget.NoSelection)
        self.stats_top_productos.setEditTriggers(QTableWidget.NoEditTriggers)
        self.stats_top_productos.setAlternatingRowColors(True)

        top_layout.addWidget(self.stats_top_productos)
        lay.addWidget(top_group)

        lay.addStretch()

        return scroll

    def _crear_kpi_card(self, title, value, color):
        card = QWidget()
        card.setStyleSheet(f"""
            QWidget {{
                background-color: white;
                border: 2px solid {color};
                border-radius: 8px;
                padding: 16px;
            }}
        """)
        layout = QVBoxLayout(card)

        lbl_title = QLabel(title)
        lbl_title.setStyleSheet(f"color: {color}; font-size: 11px; font-weight: bold;")
        lbl_title.setAlignment(Qt.AlignCenter)

        lbl_value = QLabel(value)
        lbl_value.setStyleSheet(f"color: {color}; font-size: 24px; font-weight: bold;")
        lbl_value.setAlignment(Qt.AlignCenter)
        lbl_value.setObjectName("kpi_value")

        layout.addWidget(lbl_title)
        layout.addWidget(lbl_value)

        return card

    def _obtener_ventas_rango(self, dt_min, dt_max, sucursal=None):
        """Obtiene ventas en un rango de fechas usando listar_por_fecha"""
        ventas = []
        fecha_actual = dt_min.date()
        fecha_fin = dt_max.date()

        while fecha_actual < fecha_fin:
            ventas_dia = self.repo.listar_por_fecha(fecha_actual, sucursal)
            ventas.extend(ventas_dia)
            fecha_actual += timedelta(days=1)

        return ventas

    def _actualizar_estadisticas(self):
        """Actualiza las estadísticas basándose en los filtros actuales"""
        from collections import defaultdict

        # Obtener filtros del tab Listado
        dt_min, dt_max = self._rango_fechas()
        suc = self.cmb_sucursal.currentData()
        forma_txt = self.cmb_forma.currentText().lower()

        # Actualizar banner de filtros
        sucursal_nombre = self.cmb_sucursal.currentText()
        forma_nombre = self.cmb_forma.currentText()
        filtros_texto = f"📊 Mostrando estadísticas: {dt_min.date().strftime('%d/%m/%Y')} - {dt_max.date().strftime('%d/%m/%Y')}"
        filtros_texto += f"  |  Sucursal: {sucursal_nombre}"
        filtros_texto += f"  |  Forma de pago: {forma_nombre}"
        self.stats_filtros_banner.setText(filtros_texto)

        # Consultar ventas usando el método correcto
        ventas = self._obtener_ventas_rango(dt_min, dt_max, suc)
        print(f"[DEBUG] Total ventas encontradas: {len(ventas)}")

        # Filtrar por forma de pago si es necesario
        if forma_txt in ("efectivo", "tarjeta"):
            ventas = [v for v in ventas if v.modo_pago.lower() == forma_txt]
            print(f"[DEBUG] Ventas después de filtrar por {forma_txt}: {len(ventas)}")

        # Calcular KPIs
        total = sum(getattr(v, 'total', 0) or 0 for v in ventas)
        cantidad = len(ventas)
        promedio = total / cantidad if cantidad > 0 else 0
        total_interes = sum(getattr(v, 'interes_monto', 0) or 0 for v in ventas)
        print(f"[DEBUG] KPIs - Total: ${total}, Cantidad: {cantidad}, Promedio: ${promedio}")

        # Actualizar KPI cards
        self.kpi_total.findChild(QLabel, "kpi_value").setText(f"${total:,.2f}")
        self.kpi_cantidad.findChild(QLabel, "kpi_value").setText(str(cantidad))
        self.kpi_promedio.findChild(QLabel, "kpi_value").setText(f"${promedio:,.2f}")
        self.kpi_interes.findChild(QLabel, "kpi_value").setText(f"${total_interes:,.2f}")

        # Generar gráfico de barras
        self._generar_grafico_ventas(ventas, dt_min, dt_max)

        # Generar gráfico de torta para formas de pago
        self._generar_grafico_formas_pago(ventas)

        # Mostrar comparativa solo si se seleccionó "Todas" las sucursales
        if suc is None:  # "Todas"
            self.stats_comparativa_group.setVisible(True)
            self._generar_comparativa_sucursales(dt_min, dt_max, forma_txt)
        else:
            self.stats_comparativa_group.setVisible(False)

        # Calcular top productos
        self._calcular_top_productos(ventas)

    def _generar_grafico_ventas(self, ventas, dt_min, dt_max):
        """Genera un gráfico de barras con las ventas por día"""
        try:
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure
            from collections import defaultdict

            print(f"[DEBUG] Generando gráfico de ventas con {len(ventas)} ventas")

            # Limpiar contenedor anterior
            for i in reversed(range(self.stats_chart_layout.count())):
                widget = self.stats_chart_layout.itemAt(i).widget()
                if widget:
                    widget.deleteLater()

            # Agrupar ventas por día
            ventas_por_dia = defaultdict(float)
            for v in ventas:
                fecha = getattr(v, 'fecha', None)
                if fecha:
                    dia = fecha.date()
                    total = getattr(v, 'total', 0) or 0
                    ventas_por_dia[dia] += total

            # Ordenar por fecha
            dias = sorted(ventas_por_dia.keys())
            totales = [ventas_por_dia[d] for d in dias]

            print(f"[DEBUG] Días con ventas: {len(dias)}")

            # Crear figura
            fig = Figure(figsize=(10, 4), dpi=100)
            ax = fig.add_subplot(111)

            if dias:
                bars = ax.bar(range(len(dias)), totales, color='#2e7d32')
                ax.set_xticks(range(len(dias)))
                ax.set_xticklabels([d.strftime('%d/%m') for d in dias], rotation=45, ha='right')
                ax.set_ylabel('Total ($)')
                ax.set_title('Ventas por Día', fontsize=14, fontweight='bold')
                ax.grid(axis='y', alpha=0.3)

                # Añadir valores encima de las barras
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'${height:,.0f}',
                           ha='center', va='bottom', fontsize=8)
            else:
                ax.text(0.5, 0.5, 'No hay datos para mostrar',
                       ha='center', va='center', transform=ax.transAxes,
                       fontsize=14, color='gray')

            fig.tight_layout()

            # Agregar canvas al layout
            canvas = FigureCanvas(fig)
            canvas.setMinimumHeight(300)  # Asegurar altura mínima
            self.stats_chart_layout.addWidget(canvas)
            canvas.draw()  # Forzar renderizado
            print("[DEBUG] Gráfico de barras agregado al layout")

        except Exception as e:
            import traceback
            print(f"Error generando gráfico: {e}")
            traceback.print_exc()

    def _generar_grafico_formas_pago(self, ventas):
        """Genera un gráfico de torta mostrando la distribución de formas de pago"""
        try:
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure
            from collections import defaultdict

            print(f"[DEBUG] Generando gráfico de torta con {len(ventas)} ventas")

            # Limpiar contenedor anterior
            for i in reversed(range(self.stats_pie_layout.count())):
                widget = self.stats_pie_layout.itemAt(i).widget()
                if widget:
                    widget.deleteLater()

            # Agrupar ventas por forma de pago
            formas_pago = defaultdict(float)
            for v in ventas:
                modo = getattr(v, 'modo_pago', 'Desconocido') or 'Desconocido'
                total = getattr(v, 'total', 0) or 0
                formas_pago[modo] += total

            print(f"[DEBUG] Formas de pago encontradas: {dict(formas_pago)}")

            if not formas_pago:
                print("[DEBUG] No hay formas de pago para mostrar")
                return

            # Crear figura
            fig = Figure(figsize=(8, 5), dpi=100)
            ax = fig.add_subplot(111)

            # Datos para el gráfico
            labels = list(formas_pago.keys())
            sizes = list(formas_pago.values())
            colors = ['#2e7d32', '#1976d2', '#f57c00', '#c62828'][:len(labels)]
            explode = [0.05] * len(labels)  # Separar ligeramente todas las porciones

            # Crear gráfico de torta
            wedges, texts, autotexts = ax.pie(
                sizes,
                labels=labels,
                autopct='%1.1f%%',
                startangle=90,
                colors=colors,
                explode=explode
            )

            # Mejorar apariencia del texto
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(10)

            for text in texts:
                text.set_fontsize(11)
                text.set_fontweight('bold')

            ax.set_title('Distribución por Forma de Pago', fontsize=14, fontweight='bold')
            ax.axis('equal')  # Para que sea circular

            fig.tight_layout()

            # Agregar canvas al layout
            canvas = FigureCanvas(fig)
            canvas.setMinimumHeight(350)  # Asegurar altura mínima
            self.stats_pie_layout.addWidget(canvas)
            canvas.draw()  # Forzar renderizado
            print("[DEBUG] Gráfico de torta agregado al layout")

        except Exception as e:
            import traceback
            print(f"Error generando gráfico de torta: {e}")
            traceback.print_exc()

    def _generar_comparativa_sucursales(self, dt_min, dt_max, forma_txt):
        """Genera gráficos comparativos entre sucursales"""
        try:
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure

            # Limpiar contenedor anterior
            for i in reversed(range(self.stats_comparativa_layout.count())):
                widget = self.stats_comparativa_layout.itemAt(i).widget()
                if widget:
                    widget.deleteLater()

            # Obtener datos por sucursal
            sucursales_data = {}
            for sucursal in ["Sarmiento", "Salta"]:
                ventas = self._obtener_ventas_rango(dt_min, dt_max, sucursal=sucursal)
                if forma_txt in ("efectivo", "tarjeta"):
                    ventas = [v for v in ventas if v.modo_pago.lower() == forma_txt]

                total = sum(getattr(v, 'total', 0) or 0 for v in ventas)
                cantidad = len(ventas)
                promedio = total / cantidad if cantidad > 0 else 0

                sucursales_data[sucursal] = {
                    'total': total,
                    'cantidad': cantidad,
                    'promedio': promedio
                }

            # Crear figura con 2 subplots
            fig = Figure(figsize=(12, 4), dpi=100)

            # Subplot 1: Total por sucursal
            ax1 = fig.add_subplot(121)
            sucursales = list(sucursales_data.keys())
            totales = [sucursales_data[s]['total'] for s in sucursales]
            bars1 = ax1.bar(sucursales, totales, color=['#2e7d32', '#1976d2'])
            ax1.set_ylabel('Total ($)')
            ax1.set_title('Total Vendido por Sucursal')
            ax1.grid(axis='y', alpha=0.3)

            for bar in bars1:
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height,
                        f'${height:,.0f}',
                        ha='center', va='bottom')

            # Subplot 2: Cantidad de ventas por sucursal
            ax2 = fig.add_subplot(122)
            cantidades = [sucursales_data[s]['cantidad'] for s in sucursales]
            bars2 = ax2.bar(sucursales, cantidades, color=['#f57c00', '#c62828'])
            ax2.set_ylabel('Cantidad')
            ax2.set_title('Cantidad de Ventas por Sucursal')
            ax2.grid(axis='y', alpha=0.3)

            for bar in bars2:
                height = bar.get_height()
                ax2.text(bar.get_x() + bar.get_width()/2., height,
                        f'{int(height)}',
                        ha='center', va='bottom')

            fig.tight_layout()

            # Agregar canvas con altura mínima
            canvas = FigureCanvas(fig)
            canvas.setMinimumHeight(300)
            canvas.setMaximumHeight(400)
            self.stats_comparativa_layout.addWidget(canvas)

            # Agregar separador con espacio
            self.stats_comparativa_layout.addSpacing(20)

            # Agregar tabla resumen con separador visual
            lbl_tabla = QLabel("Resumen Comparativo:")
            lbl_tabla.setStyleSheet("font-weight: bold; font-size: 12px; margin-top: 15px; margin-bottom: 5px;")
            self.stats_comparativa_layout.addWidget(lbl_tabla)

            tabla = QTableWidget(len(sucursales), 3)
            tabla.setHorizontalHeaderLabels(["Sucursal", "Total", "Promedio"])

            # Calcular altura basada en filas
            row_height = 30
            header_height = 30
            total_height = (row_height * len(sucursales)) + header_height + 10
            tabla.setMinimumHeight(total_height)
            tabla.setMaximumHeight(total_height + 20)

            tabla.verticalHeader().setVisible(False)
            tabla.horizontalHeader().setStretchLastSection(True)
            tabla.setSelectionMode(QTableWidget.NoSelection)
            tabla.setEditTriggers(QTableWidget.NoEditTriggers)

            for i, suc in enumerate(sucursales):
                tabla.setItem(i, 0, QTableWidgetItem(suc))
                tabla.setItem(i, 1, QTableWidgetItem(f"${sucursales_data[suc]['total']:,.2f}"))
                tabla.setItem(i, 2, QTableWidgetItem(f"${sucursales_data[suc]['promedio']:,.2f}"))
                tabla.setRowHeight(i, row_height)

            self.stats_comparativa_layout.addWidget(tabla)
            self.stats_comparativa_layout.addStretch()  # Agregar stretch al final

        except Exception as e:
            print(f"Error generando comparativa: {e}")

    def _calcular_top_productos(self, ventas):
        """Calcula y muestra los top 10 productos más vendidos"""
        from collections import defaultdict

        productos_stats = defaultdict(lambda: {'cantidad': 0, 'total': 0.0})

        for venta in ventas:
            items = getattr(venta, 'items', []) or []
            for item in items:
                prod = getattr(item, 'producto', None)
                if prod:
                    nombre = getattr(prod, 'nombre', 'Sin nombre')
                    cantidad = getattr(item, 'cantidad', 0) or 0
                    precio = getattr(item, 'precio_unit', 0) or 0

                    productos_stats[nombre]['cantidad'] += cantidad
                    productos_stats[nombre]['total'] += cantidad * precio

        # Ordenar por cantidad y tomar top 10
        top_productos = sorted(productos_stats.items(),
                              key=lambda x: x[1]['cantidad'],
                              reverse=True)[:10]

        # Actualizar tabla con altura ajustada
        self.stats_top_productos.setRowCount(len(top_productos))

        # Altura fija por fila
        row_height = 35
        for i, (nombre, stats) in enumerate(top_productos):
            self.stats_top_productos.setItem(i, 0, QTableWidgetItem(nombre))
            self.stats_top_productos.setItem(i, 1, QTableWidgetItem(str(int(stats['cantidad']))))
            self.stats_top_productos.setItem(i, 2, QTableWidgetItem(f"${stats['total']:,.2f}"))
            self.stats_top_productos.setRowHeight(i, row_height)

        # Ajustar altura según contenido
        row_count = self.stats_top_productos.rowCount()
        if row_count > 0:
            header_height = self.stats_top_productos.horizontalHeader().height()
            total_height = (row_height * row_count) + header_height + 20
            self.stats_top_productos.setMinimumHeight(total_height)
            self.stats_top_productos.setMaximumHeight(total_height + 30)
        else:
            # Si no hay productos, altura mínima
            self.stats_top_productos.setMinimumHeight(100)
            self.stats_top_productos.setMaximumHeight(100)

    def recargar_historial(self):
    # Alias compatible con llamadas existentes desde main_window
        self.refrescar()

    # ------------------- Carga / filtros -------------------
    def _rango_fechas(self):
        d1 = self.dt_desde.date().toPyDate()
        d2 = self.dt_hasta.date().toPyDate()
        # hasta fin del día
        dt_min = datetime.combine(d1, dtime.min)
        dt_max = datetime.combine(d2 + timedelta(days=1), dtime.min)
        return dt_min, dt_max

    def refrescar(self):
        dt_min, dt_max = self._rango_fechas()
        suc = self.cmb_sucursal.currentData()
        forma_txt = self.cmb_forma.currentText().lower()

        # Del repo si existe listar_por_rango; si no, fallback por fecha día a día
        ventas = []
        try:
            if hasattr(self.repo, "listar_por_rango"):
                ventas = self.repo.listar_por_rango(dt_min, dt_max, sucursal=suc)   # si tu repo soporta sucursal
            else:
                # fallback: día a día
                cur = dt_min.date()
                while cur < dt_max.date():
                    try:
                        ventas += self.repo.listar_por_fecha(cur, sucursal=suc)
                    except Exception:
                        ventas += self.repo.listar_por_fecha(cur)
                    cur += timedelta(days=1)
        except Exception:
            ventas = []

        # Filtro por forma
        if forma_txt != "todas":
            def _forma(v):
                raw = (getattr(v, "forma_pago", "") or getattr(v, "modo_pago", "") or getattr(v, "modo", "") or "").lower()
                return "tarjeta" if raw.startswith("tarj") else "efectivo"
            ventas = [v for v in ventas if _forma(v) == forma_txt]

        # Filtro por texto
        q = (self.txt_buscar.text() or "").strip().lower()
        if q:
            def _match(v):
                nro = str(getattr(v, "numero_ticket", "") or getattr(v, "id", ""))
                return q in nro.lower()
            ventas = [v for v in ventas if _match(v)]

        # Rellenar tabla
        self._ventas_cache = ventas
        self._pintar_tabla()

    def _pintar_tabla(self):
        # Asegurar 15 columnas y headers (por si no se hizo en __init__)
        if self.tbl.columnCount() != 15:
            self.tbl.setColumnCount(15)
            self.tbl.setHorizontalHeaderLabels([
                "Nº Ticket", "Fecha/Hora", "Sucursal", "Forma Pago",
                "Cuotas", "Interés", "Descuento", "Monto x cuota",
                "Total", "Pagado", "Vuelto", "CAE", "Vto CAE", "Comentario", "ID"
            ])

        self.tbl.setRowCount(0)

        total = 0.0
        tot_efectivo = 0.0
        tot_tarjeta  = 0.0

        for v in self._ventas_cache:
            row = self.tbl.rowCount()
            self.tbl.insertRow(row)

            # Campos base
            nro = str(getattr(v, "numero_ticket", "") or getattr(v, "id", ""))
            fch = getattr(v, "fecha", None)
            hora = fch.strftime("%Y-%m-%d %H:%M") if fch else ""
            suc = getattr(v, "sucursal", "") or ""

            forma_raw = (getattr(v, "forma_pago", "") or getattr(v, "modo_pago", "") or getattr(v, "modo", "") or "").lower()
            forma = "Tarjeta" if forma_raw.startswith("tarj") else "Efectivo"

            try:
                cuotas = int(getattr(v, "cuotas", 0) or 0)
            except Exception:
                cuotas = 0

            try:
                tot = float(getattr(v, "total", 0.0) or 0.0)
            except Exception:
                tot = 0.0

            total += tot
            if forma == "Tarjeta":
                tot_tarjeta += tot
            else:
                tot_efectivo += tot

            try:
                interes = float(_get_any(v, ["interes_monto", "interes", "monto_interes"], 0.0) or 0.0)
            except Exception:
                interes = 0.0
            try:
                descuento = float(_get_any(v, ["descuento_monto", "descuento", "monto_descuento"], 0.0) or 0.0)
            except Exception:
                descuento = 0.0
                
            monto_cuota = (tot / cuotas) if (forma == "Tarjeta" and cuotas) else 0.0
            
            

            # Pagado / Vuelto (solo efectivo)
            pagado_txt = "-"
            vuelto_txt = "-"
            if forma == "Efectivo":
                try:
                    pv = getattr(v, "pagado", None)
                    vv = getattr(v, "vuelto", None)
                    if pv is not None:
                        pagado_txt = f"${float(pv):.2f}"
                    if vv is not None:
                        vuelto_txt = f"${float(vv):.2f}"
                except Exception:
                    pass

            # Comentario
            coment = self._obtener_comentario(v)

            # Campos AFIP
            cae = getattr(v, "afip_cae", None) or "-"
            cae_vto = getattr(v, "afip_cae_vencimiento", None) or "-"

            # Orden EXACTO de columnas (coincide con headers)
            data = [
                nro,                                  # Nº Ticket
                hora,                                 # Fecha/Hora
                suc,                                  # Sucursal
                forma,                                # Forma Pago
                (str(cuotas) if cuotas else "-"),     # Cuotas
                (f"${interes:.2f}" if interes else "-"),      # Interés
                (f"${descuento:.2f}" if descuento else "-"),  # Descuento
                (f"${monto_cuota:.2f}" if monto_cuota else "-"),  # Monto x cuota
                f"${tot:.2f}",                        # Total
                pagado_txt,                           # Pagado
                vuelto_txt,                           # Vuelto
                str(cae),                             # CAE
                str(cae_vto),                         # Vto CAE
                coment,                               # Comentario
                str(getattr(v, "id", ""))             # ID
            ]

            for c, val in enumerate(data):
                it = QTableWidgetItem(val)
                it.setTextAlignment(Qt.AlignCenter)
                self.tbl.setItem(row, c, it)

        # Resumen inferior
        self.lbl_resumen.setText(
            f"{len(self._ventas_cache)} ventas — Efectivo ${tot_efectivo:.2f} — Tarjeta ${tot_tarjeta:.2f} — Total ${total:.2f}"
        )

        # Ocultar ID (última columna)
        self.tbl.setColumnHidden(14, True)

    def _obtener_comentario(self, v) -> str:
        # 1) Atributos directos de la venta
        for k in ("comentario", "motivo", "nota"):
            val = getattr(v, k, None)
            if val:
                return str(val)

        # 2) Logs del repositorio (nombres alternativos)
        vid = getattr(v, "id", None)
        if not vid:
            return ""
        for fn in ("obtener_ultimo_log", "ult_log"):
            try:
                if hasattr(self.repo, fn):
                    ult = getattr(self.repo, fn)(vid)
                    if ult:
                        # objeto / dict / string
                        for k in ("texto", "comentario", "motivo", "nota"):
                            if hasattr(ult, k):
                                val = getattr(ult, k)
                                if val:
                                    return str(val)
                        if isinstance(ult, dict):
                            for k in ("texto", "comentario", "motivo", "nota"):
                                if ult.get(k):
                                    return str(ult[k])
                        return str(ult)
            except Exception:
                pass

        # 3) Lista de logs completa (si existe)
        try:
            if hasattr(self.repo, "listar_logs"):
                logs = self.repo.listar_logs(vid) or []
                if logs:
                    ult = logs[-1]
                    for k in ("texto", "comentario", "motivo", "nota"):
                        if hasattr(ult, k):
                            val = getattr(ult, k)
                            if val:
                                return str(val)
                    if isinstance(ult, dict):
                        for k in ("texto", "comentario", "motivo", "nota"):
                            if ult.get(k):
                                return str(ult[k])
                    return str(ult)
        except Exception:
            pass

        # 4) Fallback directo a la DB (si está el modelo)
        try:
            from app.models import VentaLog
            q = self.session.query(VentaLog).filter_by(venta_id=vid)
            try:
                q = q.order_by(VentaLog.id.desc())
            except Exception:
                pass
            log = q.first()
            if log:
                for k in ("texto", "comentario", "motivo", "nota", "mensaje"):
                    if hasattr(log, k):
                        val = getattr(log, k)
                        if val:
                            return str(val)
        except Exception:
            pass

        return ""
    # ------------------- Exportar / enviar -------------------
    def _armar_dataframe(self) -> pd.DataFrame:
        rows = []
        for v in self._ventas_cache:
            forma_raw = (getattr(v, "forma_pago", "") or getattr(v, "modo_pago", "") or getattr(v, "modo", "") or "").lower()
            forma = "Tarjeta" if forma_raw.startswith("tarj") else "Efectivo"
            rows.append({
                "ticket":   getattr(v, "numero_ticket", None) or getattr(v, "id", None),
                "fecha":    getattr(v, "fecha", None),
                "sucursal": getattr(v, "sucursal", "") or "",
                "forma":    forma,
                "cuotas":   int(getattr(v, "cuotas", 0) or 0),
                "total":    float(getattr(v, "total", 0.0) or 0.0),
                "pagado":   float(getattr(v, "pagado", 0.0) or 0.0),
                "vuelto":   float(getattr(v, "vuelto", 0.0) or 0.0),
                "comentarios": self._obtener_comentario(v),
                "interes":   float(_get_any(v, ["interes_monto", "interes", "monto_interes"], 0.0) or 0.0),
                "descuento": float(_get_any(v, ["descuento_monto", "descuento", "monto_descuento"], 0.0) or 0.0),
                "monto_cuota": (
                    (float(getattr(v, "total", 0.0) or 0.0) / int(getattr(v, "cuotas", 0) or 0))
                    if (str((getattr(v, "forma_pago", "") or getattr(v, "modo_pago", "") or getattr(v, "modo", "")).lower()).startswith("tarj")
                        and int(getattr(v, "cuotas", 0) or 0) > 0)
                    else 0.0
                ),
            })
        return pd.DataFrame(rows)


    def _autofit_sheet(ws, df, engine_name: str):
        # Calcula un ancho aproximado según el contenido
        def _maxlen(series):
            try:
                return max([len(str(series.name))] + [len(str(x)) for x in series.astype(str).tolist()])
            except Exception:
                return 12
        if "openpyxl" in (engine_name or "").lower():
            from openpyxl.utils import get_column_letter
            for i, col in enumerate(df.columns, start=1):
                width = min(max(10, _maxlen(df[col]) + 2), 60)
                ws.column_dimensions[get_column_letter(i)].width = width
        else:  # xlsxwriter
            for i, col in enumerate(df.columns):
                width = min(max(10, _maxlen(df[col]) + 2), 60)
                try:
                    ws.set_column(i, i, width)
                except Exception:
                    pass

    



    # historialventas.py
    def _armar_dataframe_items(self) -> pd.DataFrame:
        items_rows = []
        for v in self._ventas_cache:
            vid = getattr(v, "id", None)
            if not vid:
                continue

            # 1) Intento por el repo
            try:
                items = self.repo.listar_items(vid)
            except Exception:
                items = []

            # 2) Fallback directo a la DB si el repo no trae nada
            if not items:
                try:
                    from app.models import VentaItem
                    items = self.session.query(VentaItem).filter_by(venta_id=vid).all()
                except Exception:
                    items = []

            # 3) Normalización por ítem
            for it in items:
                prod = getattr(it, "producto", None)

                # Categoría (si existe)
                cat = getattr(prod, "categoria", "") if prod else ""

                # Código (varios alias) + fallback a producto
                codigo = (
                    getattr(it, "codigo", None)
                    or getattr(it, "codigo_barra", None)
                    or getattr(it, "producto_codigo", None)
                )
                if not codigo and prod:
                    codigo = getattr(prod, "codigo_barra", "") or getattr(prod, "codigo", "")

                # Nombre (varios alias) + fallback a producto
                nombre = getattr(it, "nombre", None)
                if not nombre and prod:
                    nombre = getattr(prod, "nombre", "") or getattr(prod, "descripcion", "")

                # Cantidad
                try:
                    cant = float(getattr(it, "cantidad", 1) or 1)
                except Exception:
                    cant = 1.0

                # Precio unitario (varios alias)
                try:
                    pu = float(
                        getattr(it, "precio_unitario",
                            getattr(it, "precio_unit",
                                getattr(it, "precio", 0.0)
                            )
                        ) or 0.0
                    )
                except Exception:
                    pu = 0.0

                items_rows.append({
                    "venta_id": vid,
                    "ticket": getattr(v, "numero_ticket", None) or vid,
                    "codigo": codigo or "",
                    "nombre": nombre or "",
                    "categoria": cat,
                    "cantidad": cant,
                    "precio_unitario": pu,
                    "total_linea": cant * pu
                })

        return pd.DataFrame(items_rows)


    def _exportar_a_xlsx_local(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel", "historial.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return

        # Asegurar extensión .xlsx
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            out_path = self._crear_excel(path)
            QMessageBox.information(self, "Exportar", f"Guardado:\n{out_path}")
        except Exception as e:
            QMessageBox.warning(
                self, "Exportar",
                "No se pudo crear el archivo XLSX.\n\n"
                f"Detalle: {e}\n\n"
                "Probá actualizar dependencias:\n"
                "  pip install -U pandas openpyxl XlsxWriter"
            )
    
    def _crear_excel(self, path: str, for_freq: Optional[str] = None) -> str:
        import pandas as pd, os
        df  = self._armar_dataframe()

        # Si el caller pide contenido específico (programación), consultamos config
        cfg = load_config()
        rep = (cfg.get("reports") or {}).get("historial") or {}
        excel_content = rep.get("excel_content") or rep.get("export_content") or {}

        # Mantener el checkbox para items cuando es envío manual;
        # para auto (for_freq) también respetamos si el usuario lo tilda.
        inc = self.chk_incluir_items.isChecked()
        dfi = self._armar_dataframe_items() if inc else None

        # Writer
        xw = _make_writer(path)
        if xw is None:
            base, _ = os.path.splitext(path)
            csv1 = base + ".csv"
            df.to_csv(csv1, index=False, encoding="utf-8-sig")
            if dfi is not None and not dfi.empty:
                dfi.to_csv(base + "_items.csv", index=False, encoding="utf-8-sig")
            try:
                QMessageBox.warning(
                    self, "Exportar",
                    "No se encontró un engine XLSX (openpyxl/xlsxwriter). Guardé CSV.\n"
                    "Para XLSX instalá (o actualizá):\n"
                    "  pip install -U openpyxl\n  o bien:\n  pip install -U XlsxWriter"
                )
            except Exception:
                pass
            return csv1

        try:
            with xw:
                # --- Hoja Ventas (siempre) ---
                df.to_excel(xw, index=False, sheet_name="Ventas")

                # Ajustes de ancho
                ws = None
                try:
                    ws = xw.sheets.get("Ventas") if hasattr(xw, "sheets") else None
                except Exception:
                    ws = None
                if ws is None:
                    try:
                        wb = getattr(xw, "book", None)
                        if wb is not None:
                            ws = wb["Ventas"]
                    except Exception:
                        ws = None

                _engine = (getattr(xw, "engine", "") or getattr(xw, "_engine", "")).lower()
                if ws is not None:
                    _autofit_sheet(ws, df, _engine)

                # --- Resumen al pie ---
                desde = self.dt_desde.date().toString("yyyy-MM-dd")
                hasta = self.dt_hasta.date().toString("yyyy-MM-dd")
                n = len(df.index)
                if not df.empty and "forma" in df.columns:
                    g = df.groupby("forma", dropna=False)["total"].sum()
                    tot_eff = float(g.get("Efectivo", 0.0))
                    tot_tar = float(g.get("Tarjeta",  0.0))
                    tot_all = float(df["total"].sum())
                else:
                    tot_eff = tot_tar = tot_all = 0.0

                resumen = pd.DataFrame([
                    [f"Total vendido entre {desde} y {hasta}", ""],
                    ["Efectivo", f"{tot_eff:.2f}"],
                    ["Tarjeta",  f"{tot_tar:.2f}"],
                    ["TOTAL",    f"{tot_all:.2f}"],
                ], columns=["Detalle", "Monto"])
                resumen.to_excel(xw, index=False, header=True, sheet_name="Ventas", startrow=n + 2)
                if ws is not None:
                    try:
                        _autofit_sheet(ws, resumen, _engine)
                    except Exception:
                        pass

                # --- Hojas adicionales según contenido configurado ---
                # Mapear "Diario/Semanal/Mensual" a llaves (aceptamos también español).
                freq_key = None
                if for_freq:
                    f = (for_freq or "").strip().lower()
                    if "diar" in f or f == "daily":
                        freq_key = "daily"
                    elif "seman" in f or f == "weekly":
                        freq_key = "weekly"
                    elif "mensu" in f or f == "monthly":
                        freq_key = "monthly"

                requested = (excel_content.get(freq_key) if freq_key else None) or []

                # Productos Vendidos (si se pidió por config o si el usuario marcó incluir items)
                need_products = ("productos_mas_vendidos" in requested) or (inc and dfi is not None and not dfi.empty)
                if need_products and dfi is not None and not dfi.empty:
                    cols = [c for c in ("categoria", "codigo", "nombre", "cantidad", "total_linea") if c in dfi.columns]
                    agg = dfi[cols].copy()
                    if "categoria" not in agg.columns:
                        agg["categoria"] = ""
                    prod = (agg.groupby(["categoria", "codigo", "nombre"], dropna=False)
                            .agg(cantidad=("cantidad", "sum"), monto=("total_linea", "sum"))
                            .reset_index()
                            .sort_values(["categoria", "nombre"]))
                    prod.to_excel(xw, index=False, sheet_name="Productos Vendidos")
                    try:
                        ws_prod = xw.sheets.get("Productos Vendidos")
                    except Exception:
                        ws_prod = None
                    if ws_prod is not None:
                        _autofit_sheet(ws_prod, prod, _engine)

                # (Opcional) Aquí podrías agregar otras hojas si luego definimos
                # "ventas_diarias", "resumen_semanal", "ventas_semanales" con dataframes específicos.

            # Verificación final (archivo debe existir y tener >0 bytes)
            if not os.path.exists(path) or os.path.getsize(path) <= 0:
                base, _ = os.path.splitext(path)
                csv1 = base + ".csv"
                df.to_csv(csv1, index=False, encoding="utf-8-sig")
                if dfi is not None and not dfi.empty:
                    dfi.to_csv(base + "_items.csv", index=False, encoding="utf-8-sig")
                raise IOError("El XLSX quedó vacío; se guardó CSV como respaldo.")

            return path

        except Exception as e:
            base, _ = os.path.splitext(path)
            csv1 = base + ".csv"
            try:
                df.to_csv(csv1, index=False, encoding="utf-8-sig")
                if dfi is not None and not dfi.empty:
                    dfi.to_csv(base + "_items.csv", index=False, encoding="utf-8-sig")
            finally:
                pass
            try:
                QMessageBox.warning(
                    self, "Exportar",
                    "Hubo un problema al guardar el XLSX. Guardé CSV como respaldo.\n\n"
                    f"Detalle: {e}\n\n"
                    "Sugerencia:\n  pip install -U pandas openpyxl XlsxWriter"
                )
            except Exception:
                pass
            return csv1



        


    def _exportar_y_enviar(self, for_freq: Optional[str] = None):
        cfg = load_config()
        e = (cfg.get("email") or {})

        # Destinatarios desde config
        recipients = list(filter(None, e.get("recipients") or []))
        if not recipients:
            QMessageBox.warning(self, "Correo", "Agregá al menos un destinatario en Configuración → Email.")
            return

        # Excel temporal con los filtros actuales
        tmpdir = tempfile.gettempdir()
        fname = f"historial_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fpath = os.path.join(tmpdir, fname)
        self._crear_excel(fpath, for_freq=for_freq)

        # Enviar
        try:
            subj_prefix = (e.get("subject_prefix") or "[Historial]")
            subj = f"{subj_prefix} Ventas {self.dt_desde.date().toString('yyyy-MM-dd')} a {self.dt_hasta.date().toString('yyyy-MM-dd')}"
            body = "Se adjunta el reporte con los filtros aplicados."
            _send_mail_with_attachments(subj, body, recipients, [fpath])
            QMessageBox.information(self, "Correo", "Enviado.")
        except Exception as ex:
            QMessageBox.warning(self, "Correo", f"No se pudo enviar:\n{ex}")



    # ------------------- Detalle de venta -------------------
    def _ver_detalle_venta(self, row: int, col: int):
        it = self.tbl.item(row, self.tbl.columnCount() - 1)  # ID oculto
        if not it:
            return
        try:
            venta_id = int(it.text())
        except Exception:
            return

        dlg = _VentaDetalleDialog(self.session, venta_id, self)
        dlg.exec_()


class _VentaDetalleDialog(QDialog):
    def __init__(self, session, venta_id: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Detalle venta #{venta_id}")
        self.resize(680, 400)
        lay = QVBoxLayout(self)

        tbl = QTableWidget(0, 6)
        tbl.setHorizontalHeaderLabels(["Código", "Nombre", "Cantidad", "P. Unit.", "Total línea", "Venta ID"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.verticalHeader().setVisible(False)
        lay.addWidget(tbl)

        repo = VentaRepo(session)   # :contentReference[oaicite:10]{index=10}
        try:
            items = repo.listar_items(venta_id)
        except Exception:
            items = session.query(VentaItem).filter_by(venta_id=venta_id).all()

        for it in items:
            row = tbl.rowCount()
            tbl.insertRow(row)

            # Código
            codigo = getattr(it, "codigo", None) or getattr(it, "codigo_barra", None) or ""
            if not codigo:
                prod = getattr(it, "producto", None)
                if prod:
                    codigo = getattr(prod, "codigo_barra", "") or getattr(prod, "codigo", "") or ""

            # Nombre
            nombre = getattr(it, "nombre", None)
            if not nombre:
                prod = getattr(it, "producto", None)
                if prod:
                    nombre = getattr(prod, "nombre", "") or getattr(prod, "descripcion", "")

            # Cantidad / P.Unit
            try:
                cant = float(getattr(it, "cantidad", 1) or 1)
            except Exception:
                cant = 1.0
            try:
                pu = float(
                    getattr(it, "precio_unitario",
                        getattr(it, "precio_unit",
                            getattr(it, "precio", 0.0)
                        )
                    ) or 0.0
                )
            except Exception:
                pu = 0.0

            tot_linea = cant * pu
            data = [codigo, nombre, f"{cant:.2f}", f"{pu:.2f}", f"{tot_linea:.2f}", str(venta_id)]
            for c, val in enumerate(data):
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignCenter)
                tbl.setItem(row, c, cell)

        tbl.setColumnHidden(5, True)